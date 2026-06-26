#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Name Teknik — Excel -> uygulama içe aktarma iskeleti.

Kullanım:
    python scripts/ice_aktarma.py [EXCEL_YOLU]

Varsayılan EXCEL_YOLU = MART_2026.xlsx (proje kökünde).

Bu betik ŞİMDİLİK KURU ÇALIŞIR (veritabanına dokunmaz):
  - Tüm müşteri sayfalarını + DİĞER sayfasını okur (DURUM/Sayfa6 atlanır).
  - Satırları tek şemaya (is_kaydi) normalize eder.
  - Durum / fatura / personel / müşteri değerlerini temizler (trim, boşluk,
    büyük harf, bilinen yazım düzeltmeleri).
  - Gömülü RESİM-1/RESİM-2 fotoğraflarını çıkarıp satırlara eşler.
  - Çıktı: scripts/_cikti/temiz_veri.json, ozet.json ve _cikti/fotolar/ klasörü.

Gerçek yükleme (Supabase'e yazma) ayrı bir adımda eklenecek; o adımda yalnızca
yerelde kullanılan bir SUPABASE_SERVICE_ROLE_KEY gerekir (repoya girmez).
"""
import sys
import os
import re
import json
import datetime as dt

import openpyxl

KOK = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CIKTI = os.path.join(KOK, "scripts", "_cikti")
FOTO_DIZIN = os.path.join(CIKTI, "fotolar")

ATLA = {"DURUM", "SAYFA6"}

# Excel başlığı (normalize) -> şema alanı
BASLIK_ESLEME = {
    "URUNUN ADI VEYA KODU": "cihaz_adi",
    "KARTIN ADI": "cihaz_adi",
    "FIRMA ADI": "musteri",
    "GELIS TARIHI": "gelis_tarihi",
    "CIKIS TARIHI": "cikis_tarihi",
    "DURUM BILGISI": "durum",
    "DURUM": "durum",
    "TEKNIK PERSONEL": "teknik_personel",
    "SONUC": "fatura_durumu",
    "FATURA": "fatura_durumu",
    "TEKNIK SERVIS NO": "servis_no",
    "FIS NO": "servis_no",
    "ACIKLAMA": "aciklama",
    "ACIKLAMA / SERI NO": "aciklama",
    "ILGILI KISI": "ilgili_kisi",
    "FIYAT TEKLIFI": "fiyat_teklifi",
    "FATURA BIRIM TUTARI": "fatura_tutari",
    "RESIM-1": "resim",
    "RESIM-2": "resim",
}

# Bilinen yazım/normalize düzeltmeleri (genişletilebilir)
DURUM_DUZELT = {
    "ONARILDI": "ONARILDI",
    "ONARIMDA": "ONARIMDA",
    "BAKILMADI": "BAKILMADI",
    "GERI GELEN": "GERİ GELEN",
    "IADE": "İADE",
    "İADE": "İADE",
    "SATIS": "SATIŞ",
    "SATIŞ": "SATIŞ",
}


def norm_baslik(s):
    """Başlığı eşleme için sadeleştir: büyük harf, Türkçe -> ASCII, tek boşluk."""
    if s is None:
        return ""
    s = str(s).strip()
    cevrim = str.maketrans("çÇğĞıİiöÖşŞüÜ", "cCgGiIioOsSuU")
    s = s.translate(cevrim).upper()
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def temizle(s):
    if s is None:
        return None
    s = re.sub(r"\s+", " ", str(s)).strip()
    return s or None


def buyuk(s):
    """Türkçe-duyarlı büyük harf (i -> İ, ı korunur)."""
    if s is None:
        return None
    s = temizle(s)
    if s is None:
        return None
    return s.replace("i", "İ").replace("ı", "I").upper()


def tarih_cevir(v):
    if v is None or v == "":
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for kal in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(s, kal).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None  # çözülemeyen tarih


def sayi_cevir(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v)
    s = re.sub(r"[^\d,.\-]", "", s)
    if not s:
        return None
    # Türkçe biçim: 1.234,56 -> 1234.56
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def seri_no_ayikla(aciklama):
    """Açıklamadaki 'SN:...' ifadesinden seri no çıkar."""
    if not aciklama:
        return None, aciklama
    m = re.search(r"SN[:\s]*([A-Za-z0-9\-]+)", aciklama)
    if m:
        return m.group(1), aciklama
    return None, aciklama


def main():
    yol = sys.argv[1] if len(sys.argv) > 1 else os.path.join(KOK, "MART_2026.xlsx")
    if not os.path.exists(yol):
        print("HATA: dosya yok:", yol)
        sys.exit(1)

    print("Okunuyor:", yol)
    wb = openpyxl.load_workbook(yol, data_only=True)

    kayitlar = []
    musteriler, personeller, durumlar, faturalar = set(), set(), set(), set()
    atlanan_sayfa = []

    for ws in wb.worksheets:
        if norm_baslik(ws.title) in ATLA:
            atlanan_sayfa.append(ws.title)
            continue

        # Başlık satırını bul (genelde 2. satır)
        baslik_satir = None
        for r in range(1, 4):
            degerler = [c.value for c in ws[r]]
            if any(norm_baslik(d) in BASLIK_ESLEME for d in degerler):
                baslik_satir = r
                break
        if baslik_satir is None:
            atlanan_sayfa.append(ws.title + " (başlık yok)")
            continue

        # sütun index -> alan
        sutun_alan = {}
        for idx, hucre in enumerate(ws[baslik_satir]):
            alan = BASLIK_ESLEME.get(norm_baslik(hucre.value))
            if alan:
                sutun_alan.setdefault(alan, idx)  # ilk eşleşen sütun

        sayfa_musterisi = None if "musteri" in sutun_alan else temizle(ws.title)

        for r in range(baslik_satir + 1, ws.max_row + 1):
            satir = [c.value for c in ws[r]]

            def al(alan):
                i = sutun_alan.get(alan)
                return satir[i] if i is not None and i < len(satir) else None

            cihaz = temizle(al("cihaz_adi"))
            musteri = sayfa_musterisi or temizle(al("musteri"))
            if not cihaz:
                continue  # cihaz adı yoksa gerçek kayıt değil (boş/biçim satırı)

            aciklama = temizle(al("aciklama"))
            seri, aciklama = seri_no_ayikla(aciklama)

            durum = buyuk(al("durum"))
            durum = DURUM_DUZELT.get(norm_baslik(durum), durum) if durum else None
            fatura = buyuk(al("fatura_durumu"))
            personel = buyuk(al("teknik_personel"))

            kayit = {
                "kaynak_sayfa": ws.title,
                "kaynak_satir": r,
                "musteri": musteri,
                "cihaz_adi": cihaz,
                "seri_no": seri,
                "servis_no": temizle(al("servis_no")),
                "gelis_tarihi": tarih_cevir(al("gelis_tarihi")),
                "cikis_tarihi": tarih_cevir(al("cikis_tarihi")),
                "durum": durum,
                "teknik_personel": personel,
                "fatura_durumu": fatura,
                "ilgili_kisi": temizle(al("ilgili_kisi")),
                "fiyat_teklifi": sayi_cevir(al("fiyat_teklifi")),
                "fatura_tutari": sayi_cevir(al("fatura_tutari")),
                "aciklama": aciklama,
                "fotolar": [],
            }
            if musteri:
                musteriler.add(musteri)
            if personel:
                personeller.add(personel)
            if durum:
                durumlar.add(durum)
            if fatura:
                faturalar.add(fatura)
            kayitlar.append(kayit)

    # --- Fotoğraf çıkarma (satıra eşleme) ---
    os.makedirs(FOTO_DIZIN, exist_ok=True)
    foto_sayisi = 0
    # (sayfa, satir) -> kayit dizini
    konum_index = {}
    for i, k in enumerate(kayitlar):
        konum_index[(k["kaynak_sayfa"], k["kaynak_satir"])] = i

    try:
        wb_img = openpyxl.load_workbook(yol)  # tam yükleme (görseller dahil)
        for ws in wb_img.worksheets:
            if norm_baslik(ws.title) in ATLA:
                continue
            for img in getattr(ws, "_images", []):
                try:
                    excel_satir = img.anchor._from.row + 1  # 0-index -> 1-index
                except Exception:
                    continue
                # En yakın kayıt satırını bul (resim hücresi kayıt satırında)
                idx = konum_index.get((ws.title, excel_satir))
                if idx is None:
                    # bir alt/üst satıra denk gelebilir
                    for ofs in (-1, 1, -2, 2):
                        idx = konum_index.get((ws.title, excel_satir + ofs))
                        if idx is not None:
                            break
                if idx is None:
                    continue
                try:
                    veri = img._data()
                except Exception:
                    continue
                ext = (getattr(img, "format", None) or "png").lower()
                ad = "%s_%d_%d.%s" % (
                    re.sub(r"[^A-Za-z0-9]+", "_", ws.title),
                    excel_satir,
                    len(kayitlar[idx]["fotolar"]) + 1,
                    ext,
                )
                with open(os.path.join(FOTO_DIZIN, ad), "wb") as f:
                    f.write(veri)
                kayitlar[idx]["fotolar"].append(ad)
                foto_sayisi += 1
    except Exception as e:
        print("UYARI: fotoğraf çıkarma atlandı:", e)

    # --- Çıktı ---
    os.makedirs(CIKTI, exist_ok=True)
    with open(os.path.join(CIKTI, "temiz_veri.json"), "w", encoding="utf-8") as f:
        json.dump(kayitlar, f, ensure_ascii=False, indent=2)

    ozet = {
        "toplam_kayit": len(kayitlar),
        "toplam_foto": foto_sayisi,
        "atlanan_sayfalar": atlanan_sayfa,
        "musteri_sayisi": len(musteriler),
        "personeller": sorted(personeller),
        "durumlar": sorted(durumlar),
        "fatura_durumlari": sorted(faturalar),
        "musteriler": sorted(musteriler),
    }
    with open(os.path.join(CIKTI, "ozet.json"), "w", encoding="utf-8") as f:
        json.dump(ozet, f, ensure_ascii=False, indent=2)

    print("--- ÖZET ---")
    print("Kayıt:", len(kayitlar), "| Foto:", foto_sayisi)
    print("Müşteri:", len(musteriler), "| Personel:", sorted(personeller))
    print("Durumlar:", sorted(durumlar))
    print("Fatura:", sorted(faturalar))
    print("Atlanan:", atlanan_sayfa)
    print("Çıktı:", CIKTI)


if __name__ == "__main__":
    main()
